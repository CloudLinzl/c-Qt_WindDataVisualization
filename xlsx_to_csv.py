#!/usr/bin/env python3
"""
Excel to CSV converter for wind data analysis
Converts .xlsx files to .csv format
"""

import sys
import os
import csv

def xlsx_to_csv(xlsx_path):
    """
    Convert an Excel file to CSV format

    Args:
        xlsx_path: Path to the .xlsx file
    """
    if not os.path.exists(xlsx_path):
        print(f"Error: File '{xlsx_path}' not found")
        sys.exit(1)

    # Get the output CSV path (same name, different extension)
    csv_path = os.path.splitext(xlsx_path)[0] + '.csv'

    try:
        # Try using pandas first
        try:
            import pandas as pd
            df = pd.read_excel(xlsx_path)
            df.to_csv(csv_path, index=False, encoding='utf-8')
        except ImportError:
            # Fall back to openpyxl
            try:
                from openpyxl import load_workbook

                wb = load_workbook(xlsx_path, read_only=True)
                ws = wb.active

                with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow(row)

                wb.close()
            except ImportError:
                print("Error: Neither pandas nor openpyxl is installed.")
                print("Please install one of them:")
                print("  pip install pandas openpyxl")
                print("  or")
                print("  pip install openpyxl")
                sys.exit(1)

        print(f"Successfully converted '{xlsx_path}' to '{csv_path}'")
        print(f"CSV file saved at: {csv_path}")

    except Exception as e:
        print(f"Error converting file: {str(e)}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python3 xlsx_to_csv.py <path_to_xlsx_file>")
        sys.exit(1)

    xlsx_file = sys.argv[1]
    xlsx_to_csv(xlsx_file)
